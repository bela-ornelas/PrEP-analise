import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def write_to_sheet(ws, df, include_index=False):
    """
    Escreve um DataFrame em uma worksheet existente, começando de A1.
    Limpa o conteúdo anterior para evitar sobras de dados antigos.
    """
    # Deleta todas as linhas para limpar os dados antigos
    # (Mantém gráficos em outras abas, mas cuidado se houver gráficos nesta aba referenciando estas células)
    ws.delete_rows(1, ws.max_row + 1)
    
    # Escrever novos dados
    for r in dataframe_to_rows(df, index=include_index, header=True):
        ws.append(r)

def export_to_excel(output_dir, data_fechamento, metrics_dict):
    """
    Gera o arquivo Excel de monitoramento.
    Se existir 'modelo_monitoramento.xlsx' na raiz, usa como template (preservando gráficos).
    Caso contrário, gera um arquivo novo do zero.
    """
    data_dt = pd.to_datetime(data_fechamento)
    mes = data_dt.month
    ano = data_dt.year
    
    filename = f"Monitoramento_PrEP_{mes:02d}_{ano}.xlsx"
    filepath = os.path.join(output_dir, filename)
    template_path = "modelo_monitoramento.xlsx"
    
    # Preparar Dados para Exportação (Nome da Aba -> (DataFrame, Index=True/False))
    dfs_to_export = {}
    
    # 1. Construir Aba Geral
    if 'classificacoes' in metrics_dict:
        c = metrics_dict['classificacoes']
        df_geral = pd.DataFrame({
            'Categoria': ['Procuraram PrEP', 'Iniciaram PrEP', 
                          'Teve dispensação nos últimos 12 meses', 
                          'Não teve dispensação nos últimos 12 meses', 
                          'Em PrEP atualmente', 'Estão descontinuados'],
            'Total': [c.get('Procuraram_PrEP', 0), c.get('Iniciaram_PrEP', 0),
                      c['Disp_Ultimos_12m'], c['Disp_Ultimos_12m_Nao'], 
                      c['EmPrEP_Atual'], c['Descontinuados']]
        })
        dfs_to_export['Geral'] = (df_geral, False)

    # 2. Mapear outras métricas na ORDEM DESEJADA
    # Chave no dict metrics -> (Nome da Aba no Excel, Incluir Index?)
    map_keys = [
        ('disp_mes_ano', 'Disp_total', True),
        ('novos_usuarios', 'Novos usuários', True),
        ('annual_summary', 'Em PrEP por ano', False),
        ('historico', 'Em PrEP_mes_ano', False),
        ('populacoes', 'Populações (em PrEP)', False),
        ('uf_summary', 'Dados por UF', False),
        ('mun_summary', 'Mun', False)
    ]
    
    for key, sheet_name, idx_bool in map_keys:
        if key in metrics_dict:
            dfs_to_export[sheet_name] = (metrics_dict[key], idx_bool)

    # ---------------------------------------------------------
    # MODO MODELO (Template)
    # ---------------------------------------------------------
    if os.path.exists(template_path):
        print(f"Modelo encontrado! Usando '{template_path}' como base...")
        try:
            wb = load_workbook(template_path)
            
            for sheet_name, (df, idx_bool) in dfs_to_export.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Atualiza os dados preservando o resto do arquivo
                    write_to_sheet(ws, df, idx_bool)
                else:
                    # Se o modelo não tiver a aba, cria
                    ws = wb.create_sheet(sheet_name)
                    write_to_sheet(ws, df, idx_bool)
            
            wb.save(filepath)
            print(f"Excel gerado com sucesso (baseado no modelo) em: {filepath}")
            return filepath
        except Exception as e:
            print(f"Erro ao processar modelo: {e}. Tentando método padrão...")
    
    # ---------------------------------------------------------
    # MODO PADRÃO (Raw)
    # ---------------------------------------------------------
    print(f"Gerando NOVO arquivo Excel (sem modelo) em: {filepath}")
    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        for sheet_name, (df, idx_bool) in dfs_to_export.items():
            df.to_excel(writer, sheet_name=sheet_name, index=idx_bool)

    print("Excel gerado com sucesso!")
    return filepath
